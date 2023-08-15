import re
import openpyxl


def extract_js_values(filename):
    with open(filename, 'r') as js_file:
        js_content = js_file.read()

    keys_and_values = re.findall(r'(\w+):[ \t]*([^\n,]+)', js_content)
    return keys_and_values


def write_to_excel(data, excel_filename):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Ghi tên cột
    ws['A1'] = 'STT'
    ws['B1'] = 'Key'
    ws['C1'] = 'En'

    for row, (key, value) in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=row - 1)  # Cột STT
        ws.cell(row=row, column=2, value=key)  # Cột Key
        cleaned_value = value.replace("'", "")
        ws.cell(row=row, column=3, value=cleaned_value)  # Cột En

    wb.save(excel_filename)


js_filename = 'en.js'
excel_filename = 'output.xlsx'

key_value_pairs = extract_js_values(js_filename)
write_to_excel(key_value_pairs, excel_filename)
